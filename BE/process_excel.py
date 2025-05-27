# Handling Excel Files for Quotation Function in My Django Project

# To simplify the process, I handle Excel files in two formats: .xlsx and .xls
# - Using openpyxl for .xlsx files
# - Using xlrd for .xls files

# Process Overview:
# 1. Select the file - Choose and get sheet
# 2. Retrieve cell values
# 3. Ignore validation (users can input anything they want)
# 4. Process sheet and customize the value fields
# 5. Save data to the database

# Note:
# 1. Xlsx:
# - Excel has another file type that is also saved with the .xlsx extension 
# but is not an Excel Workbook. This file is called 'Strict Open XML Spreadsheet' format,
# and the openpyxl library cannot handle this case.
# - Load workbook with data_only=True to get values

# 2. Xlrd file, the value date is a float when get cell, so we need to convert it to a date format
#   if cell.ctype == xlrd.XL_CELL_DATE:
#       dt = datetime(*xlrd.xldate_as_tuple(value, datemode))



import io
import logging
from datetime import datetime

import xlrd
from openpyxl import load_workbook


logger = logging.getLogger("django")

EXCEL_ERROR_VALUES = {'#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A'}


class QuotationImportForm(forms.Form):
    excel_file = forms.FileField()
    sheet_name = forms.ChoiceField()

    def __init__(self, *args, **kwargs):
        logger.info("=== INITIALIZING QUOTATION IMPORT FORM ===")
        super().__init__(*args, **kwargs)
        self.fields['sheet_name'].choices = [('', _('Select a sheet'))]

    def clean_excel_file(self):
        excel_file = self.cleaned_data.get('excel_file')
        if not excel_file:
            raise forms.ValidationError(_("Please select a file to upload."))

        file_name = excel_file.name.lower()
        if not file_name.endswith(('.xlsx', '.xls')):
            raise forms.ValidationError(_("Please upload an Excel file (.xlsx or .xls)"))

        try:
            sheet_names = self._get_sheet_names(excel_file)
            if not sheet_names:
                raise forms.ValidationError(_("No sheets found in the Excel file"))

            self.fields['sheet_name'].choices = [('', _('Select a sheet'))] + [(name, name) for name in sheet_names]
        except Exception as e:
            raise forms.ValidationError(_("Error reading Excel file: %s") % str(e))

        return excel_file

    def _get_sheet_names(self, file):
        file_content = file.read()
        file.seek(0)
        if file.name.lower().endswith('.xlsx'):
            return self._get_xlsx_sheet_names(file_content)
        else:
            return self._get_xls_sheet_names(file_content)

    def _get_xlsx_sheet_names(self, file_content):
        excel_file = io.BytesIO(file_content)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)  # Load workbook with data_only=True to get values
        if not workbook.sheetnames:
            raise forms.ValidationError(
                _("This file appears to be in 'Strict Open XML Spreadsheet' format. "
                  "Please save it as 'Excel Workbook (*.xlsx)' instead.")
            )
        return workbook.sheetnames

    def _get_xls_sheet_names(self, file_content):
        workbook = xlrd.open_workbook(file_contents=file_content)
        return workbook.sheet_names()

    def _is_xlrd_sheet(self, sheet):
        return isinstance(sheet, xlrd.sheet.Sheet)

    def _get_cell_value(self, sheet, cell_ref):
        """
        Get a cell value from sheet.
        For xlrd sheets, cell_ref can be (row, col) tuple.
        For openpyxl sheets, cell_ref can be cell name like 'I1'.
        """
        logger.info(f"[IMPORT] _get_cell_value called with cell_ref: {cell_ref}")
        if self._is_xlrd_sheet(sheet):
            row, col = cell_ref
            cell = sheet.cell(row, col)
            value = cell.value
            datemode = sheet.book.datemode

            # If it's a real Excel date
            if cell.ctype == xlrd.XL_CELL_DATE:
                dt = datetime(*xlrd.xldate_as_tuple(value, datemode))
                return dt.strftime('%d/%m/%Y')  # returns '20/05/25'

            # If it's a float and looks like a date serial number
            if cell.ctype == xlrd.XL_CELL_NUMBER and 30000 < value < 60000:
                try:
                    dt = datetime(*xlrd.xldate_as_tuple(value, datemode))
                    return dt.strftime('%d/%m/%Y')
                except:
                    return str(value)

            # If it's a string
            return str(value).strip()
        else:
            cell = sheet[cell_ref]
            value = cell.value if cell else ''
            logger.info(f"[IMPORT] valuedatetime: {value, isinstance(value, datetime)}")

            if isinstance(value, datetime):
                try:
                    return value.strftime('%d/%m/%Y')
                except Exception:
                    return str(value or '').strip()
            return str(value or '').strip()

    def validate_template(self, sheet):
        logger.info("[IMPORT] Entered validate_template")
        try:
            # Extract required fields
            if self._is_xlrd_sheet(sheet):
                quotation_number = "".join(self._get_cell_value(sheet, (1, c)) for c in [2, 3, 6])
                company_display = "".join(self._get_cell_value(sheet, (r, 1)) for r in [4, 5])
                person_in_charge = self._get_cell_value(sheet, (6, 1))
                sales_staff_display = self._get_cell_value(sheet, (11, 17))
                price_display = self._get_cell_value(sheet, (15, 4))
                quotation_date_display = self._get_cell_value(sheet, (1, 3))
            else:
                quotation_number = "".join(self._get_cell_value(sheet, cell) for cell in ['C2', 'D2', 'G2'])
                company_display = "".join(self._get_cell_value(sheet, cell) for cell in ['B5', 'B6'])
                person_in_charge = self._get_cell_value(sheet, 'B7')
                sales_staff_display = self._get_cell_value(sheet, 'R12')
                price_display = self._get_cell_value(sheet, 'E16')
                quotation_date_display = self._get_cell_value(sheet, 'D2')
            logger.info(f"[IMPORT] Extracted fields: quotation_number={quotation_number}, company_display={company_display}, person_in_charge={person_in_charge}, sales_staff_display={sales_staff_display}, price_display={price_display}, quotation_date_display={quotation_date_display}")
            logger.info("[IMPORT] Template validation successful")
        except forms.ValidationError:
            logger.error("[IMPORT] Template validation failed")
        except Exception as e:
            logger.error(f"[IMPORT] Error validating template: {e}", extra={'error': str(e)})

    def _get_company(self, company_display, person_in_charge):
        """
        Get company information if exists, otherwise return text values.
        """
        logger.info("Getting company info", extra={'data': {
            'company_display': company_display,
            'person_in_charge': person_in_charge
        }})
        try:
            if not company_display or company_display.strip() == '':
                raise forms.ValidationError(_("Company name cannot be empty"))

            company_display = company_display.strip()
            person_in_charge = person_in_charge.strip() if person_in_charge else "N/A"

            logger.info("Company info processed", extra={'data': {
                'company_display': company_display,
                'person_in_charge': person_in_charge
            }})
            return None, company_display, person_in_charge

        except Exception as e:
            logger.error("Error handling company", extra={'data': {
                'company_display': company_display,
                'error': str(e)
            }})
            raise forms.ValidationError(
                _("Error processing company information: %s") % str(e)
            )

    def _get_sales_staff(self, sales_staff_name):
        """
        Get sales staff user profile from name.
        """
        logger.info("Getting sales staff info", extra={'data': {
            'sales_staff_name': sales_staff_name
        }})
        try:
            if not sales_staff_name or sales_staff_name.strip() == '':
                logger.warning("Empty sales staff name provided")
                return None, 'N/A'

            sales_staff_name = sales_staff_name.strip()

            # Try to find staff by name_jp or name_en
            staff = UserProfile.objects.filter(
                name_jp__iexact=sales_staff_name
            ).first() or UserProfile.objects.filter(
                name_en__iexact=sales_staff_name
            ).first()

            if staff:
                logger.info("Found sales staff", extra={'data': {
                    'staff_id': staff.id,
                    'staff_name': sales_staff_name
                }})
                return staff, sales_staff_name

            # If staff not found, log warning and return default
            logger.warning("Sales staff not found", extra={'data': {
                'sales_staff_name': sales_staff_name
            }})
            return None, sales_staff_name

        except Exception as e:
            logger.error("Error getting sales staff", extra={'data': {
                'error': str(e)
            }})
            return None, sales_staff_name

    def process_excel(self):
        logger.info("[IMPORT] Entered process_excel")
        excel_file = self.cleaned_data.get('excel_file')
        sheet_name = self.cleaned_data.get('sheet_name')
        logger.info(f"[IMPORT] Received excel_file: {getattr(excel_file, 'name', None)}, sheet_name: {sheet_name}")
        try:
            file_content = excel_file.read()
            excel_file.seek(0)  # Reset file pointer
            logger.info(f"[IMPORT] Read file content, size: {len(file_content)} bytes")
            if excel_file.name.lower().endswith('.xlsx'):
                workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                available_sheets = workbook.sheetnames
                logger.info(f"[IMPORT] XLSX available_sheets: {available_sheets}")

                def get_sheet(wb, name):
                    return wb[name]
            elif excel_file.name.lower().endswith('.xls'):
                workbook = xlrd.open_workbook(file_contents=file_content)
                available_sheets = workbook.sheet_names()
                logger.info(f"[IMPORT] XLS available_sheets: {available_sheets}")

                def get_sheet(wb, name):
                    return wb.sheet_by_name(name)
            else:
                logger.error(f"[IMPORT] Unsupported file format: {excel_file.name}")
                raise forms.ValidationError(_("Unsupported file format"))
            if sheet_name not in available_sheets:
                logger.error(f"[IMPORT] Sheet not found: {sheet_name}, available: {available_sheets}")
                raise forms.ValidationError(_("Selected sheet not found in file"))
            sheet = get_sheet(workbook, sheet_name)
            logger.info(f"[IMPORT] Got sheet: {sheet_name}")
            self.validate_template(sheet)
            result = self._process_sheet_data(sheet)
            logger.info(f"[IMPORT] Excel processing completed, result: {result}")
            return result
        except Exception as e:
            logger.error(f"[IMPORT] Error processing Excel file: {e}", extra={'error': str(e), 'file_name': getattr(excel_file, 'name', None)})
            raise forms.ValidationError(_("Error processing Excel file: %s") % str(e))

    def _process_sheet_data(self, sheet):
        logger.info("[IMPORT] Entered _process_sheet_data")
        try:
            if isinstance(sheet, xlrd.sheet.Sheet):
                quotation_number = "".join(self._get_cell_value(sheet, (1, c)) for c in [2, 3, 6])
                person_in_charge = self._get_cell_value(sheet, (6, 1))
                sales_staff_display = self._get_cell_value(sheet, (11, 17))
                price_display = self._get_cell_value(sheet, (15, 4))
                quotation_date_display = self._get_cell_value(sheet, (1, 3))

                company_display = ""
                company_name = self._get_cell_value(4, 1)
                branch_name = self._get_cell_value(5, 1)
                if company_name and branch_name:
                    company_display = f"{company_name}\n{branch_name}"
                elif branch_name:
                    company_display = branch_name
                elif company_name:
                    company_display = company_name
                else:
                    company_display = "N/A"

                price_cell = sheet.cell(15, 4)
                if price_cell.ctype == xlrd.XL_CELL_ERROR:
                    price_display = None
                else:
                    price_display = str(price_cell.value).strip()
                    if price_display in EXCEL_ERROR_VALUES or not price_display:
                        price_display = None
            else:
                quotation_number = "".join(self._get_cell_value(sheet, cell) for cell in ['C2', 'D2', 'G2'])
                person_in_charge = self._get_cell_value(sheet, 'B7')
                sales_staff_display = self._get_cell_value(sheet, 'R12')
                price_display = self._get_cell_value(sheet, 'E16')
                quotation_date_display = self._get_cell_value(sheet, 'D2')

                company_display = ""
                company_name = self._get_cell_value(sheet, 'B5')
                branch_name = self._get_cell_value(sheet, 'B6')
                if company_name and branch_name:
                    company_display = f"{company_name}\n{branch_name}"
                elif branch_name:
                    company_display = branch_name
                elif company_name:
                    company_display = company_name
                else:
                    company_display = "N/A"

                if price_display in EXCEL_ERROR_VALUES or not price_display:
                    price_display = None
            logger.info(f"[IMPORT] Sheet data extracted: quotation_number={quotation_number}, company_display={company_display}, person_in_charge={person_in_charge}, sales_staff_display={sales_staff_display}, price_display={price_display}, quotation_date_display={quotation_date_display}")
            with transaction.atomic():
                company, company_display, person_in_charge = self._get_company(
                    company_display, person_in_charge
                )
                sales_staff_user, sales_staff_display = self._get_sales_staff(
                    sales_staff_display
                )
            quotation_data = {
                'quotation_number': quotation_number,
                'company': company,
                'company_display': company_display or 'N/A',
                'person_in_charge': person_in_charge or 'N/A',
                'sales_staff': sales_staff_user,
                'sales_staff_display': sales_staff_display or 'N/A',
                'price_display': price_display or 'N/A',
                'quotation_date_display': quotation_date_display or 'N/A',
                'result': 'Not ordered'
            }
            logger.info(f"[IMPORT] Quotation data prepared: {quotation_data}")
            return [quotation_data]
        except forms.ValidationError:
            logger.error("[IMPORT] Validation error in _process_sheet_data")
        except Exception as e:
            logger.error(f"[IMPORT] Error processing sheet data: {e}", extra={'error': str(e)})
